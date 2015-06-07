__author__ = 'Carlos'

# Create .lang file from the export spreadsheet for every language in the sheet
# Use details from the sheet to maintain order and structure based off English original .lang file

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from codecs import decode, encode

# Open spreadsheet

wb = load_workbook('locDB.xlsx')
ws = wb.active

# Cell function

def cstring( a ):
    stringed = str(a)
    stringed = stringed.partition('.')[2].strip('>')
    return stringed

# Determine languages in file, create list

languages = []
crange = list(ws['C1:Z1'].next())
for lang in crange:
    if ws[cstring(lang)].value is None:
        break
    languages.append(ws[cstring(lang)].value)


# Function to take a language from the spreadsheet

def listlang(l):
    col = languages.index(l)+3
    li = []
    rang = ws.iter_rows(get_column_letter(col)+'2:'+get_column_letter(col)+'200')
    sidrang = ws.iter_rows('B2:B200')
    for s, i in zip(rang, sidrang):
        st = ws[cstring(s[0])].value
        sid = ws[cstring(i[0])].value

        if str(sid).partition('-')[0] == 'Empty':
            li.append('')

        else:
            if st is not None:
                li.append(sid + '=' + st)
    return li

# Function to generate a single .lang from result of previous function

def writelist(l):
    f = open('langex\\' + l + '.lang', 'w+')
    for line in listlang(l):
        f.write(encode(line, 'utf-8') + '\n')
    f.close()

# Do a pass of every language

for l in languages:
    writelist(l)
